from flask import Flask, request, jsonify, render_template, send_file, make_response
import pickle
import pandas as pd
import numpy as np
import os
from io import BytesIO
import matplotlib.pyplot as plt
import base64
import io

# 初始化Flask应用
app = Flask(__name__)

# 特征名称
columns = ['C', 'P', 'S', 'YS', 'UTS']

# 获取当前脚本所在目录
current_dir = os.path.dirname(os.path.abspath(__file__))

# 加载训练好的模型
try:
    with open(os.path.join(current_dir, 'rf_fatigue_model.pkl'), 'rb') as f:
        rf_fatigue = pickle.load(f)
    print("✅ 疲劳极限模型加载成功")
    
    with open(os.path.join(current_dir, 'rf_sri_model.pkl'), 'rb') as f:
        rf_sri = pickle.load(f)
    print("✅ SRI模型加载成功")
except Exception as e:
    print(f"❌ 模型加载失败: {e}")
    print(f"当前工作目录: {os.getcwd()}")
    print(f"脚本所在目录: {current_dir}")
    # 为了让应用能启动，我们创建虚拟模型
    class DummyModel:
        def predict(self, X):
            return [0.0]
    rf_fatigue = DummyModel()
    rf_sri = DummyModel()
    print("⚠️ 使用虚拟模型，请检查模型文件路径")

# 主页路由 - 新的首页
@app.route('/')
def homepage():
    return render_template('homepage.html')

# 材料类型选择页面
@app.route('/material/<material_type>')
def material_selection(material_type):
    # 材料类型映射
    material_names = {
        'steel_hcf': '钢材高周疲劳预测',
        'steel_lcf': '钢材低周疲劳预测', 
        'titanium_hcf': '钛合金高周疲劳预测',
        'titanium_lcf': '钛合金低周疲劳预测'
    }
    
    if material_type not in material_names:
        return render_template('error.html', message='不支持的材料类型'), 404
    
    # 根据材料类型直接跳转到相应页面
    if material_type == 'steel_hcf':
        # 钢材高周疲劳预测 - 显示选择页面
        return render_template('material_selection.html', 
                             material_type=material_type,
                             material_name=material_names[material_type])
    elif material_type == 'steel_lcf':
        # 钢材低周疲劳预测 - 功能开发中
        return render_template('under_development.html', 
                             material_name='钢材低周疲劳预测',
                             status_message='功能开发中')
    elif material_type in ['titanium_hcf', 'titanium_lcf']:
        # 钛合金预测 - 数据收集中
        return render_template('under_development.html', 
                             material_name=material_names[material_type],
                             status_message='数据收集中')
    else:
        return render_template('under_development.html', 
                             material_name='未知材料',
                             status_message='功能开发中')

# 单独预测页面路由
@app.route('/predict_single/<material_type>')
def predict_single_page(material_type):
    if material_type == 'steel_hcf':
        # 跳转到现有的钢材高周疲劳预测页面
        return render_template('index.html')
    elif material_type == 'steel_lcf':
        # 钢材低周疲劳预测 - 功能开发中
        return render_template('under_development.html', 
                             material_name='钢材低周疲劳预测',
                             status_message='功能开发中')
    elif material_type in ['titanium_hcf', 'titanium_lcf']:
        # 钛合金预测 - 数据收集中
        material_names = {
            'titanium_hcf': '钛合金高周疲劳预测',
            'titanium_lcf': '钛合金低周疲劳预测'
        }
        return render_template('under_development.html', 
                             material_name=material_names[material_type],
                             status_message='数据收集中')
    else:
        return render_template('under_development.html', 
                             material_name='未知材料',
                             status_message='功能开发中')

# 批量预测页面路由
@app.route('/predict_batch/<material_type>')
def predict_batch_page(material_type):
    material_names = {
        'steel_hcf': '钢材高周疲劳预测',
        'steel_lcf': '钢材低周疲劳预测',
        'titanium_hcf': '钛合金高周疲劳预测',
        'titanium_lcf': '钛合金低周疲劳预测'
    }
    
    if material_type == 'steel_hcf':
        # 钢材高周疲劳预测 - 正常功能
        return render_template('batch_upload.html',
                             material_type=material_type,
                             material_name=material_names.get(material_type, '未知材料'))
    elif material_type == 'steel_lcf':
        # 钢材低周疲劳预测 - 功能开发中
        return render_template('under_development.html', 
                             material_name='钢材低周疲劳预测',
                             status_message='功能开发中')
    elif material_type in ['titanium_hcf', 'titanium_lcf']:
        # 钛合金预测 - 数据收集中
        return render_template('under_development.html', 
                             material_name=material_names[material_type],
                             status_message='数据收集中')
    else:
        return render_template('under_development.html', 
                             material_name='未知材料',
                             status_message='功能开发中')

# 原有的单独预测接口（保持不变）
@app.route('/predict', methods=['POST'])
def predict():
    # 获取输入的五个特征
    data = request.get_json(force=True)

    # 输入特征：C, P, S, YS, UTS
    C = data['C']
    P = data['P']
    S = data['S']
    YS = data['YS']
    UTS = data['UTS']

    # 构建特征向量，并转换为 DataFrame
    input_features = pd.DataFrame([[C, P, S, YS, UTS]], columns=columns)

    # 使用Fatigue Limit模型进行预测
    fatigue_limit_pred = rf_fatigue.predict(input_features)[0]

    # 使用SRI模型进行预测
    sri_pred = rf_sri.predict(input_features)[0]

    # 返回预测结果
    return jsonify({
        'fatigue_limit': fatigue_limit_pred,
        'SRI': sri_pred
    })
# 设置支持中文的字体（适用于Windows/Linux）
try:
    plt.rcParams['font.family'] = 'SimHei'  # 黑体
except:
    # 如果SimHei不可用，尝试其他中文字体
    try:
        plt.rcParams['font.family'] = 'Microsoft YaHei'  # 微软雅黑
    except:
        # 如果都不可用，使用默认字体
        plt.rcParams['font.family'] = 'DejaVu Sans'
plt.rcParams['axes.unicode_minus'] = False  # 正确显示负号

@app.route('/sn_curve', methods=['POST'])
def sn_curve():
    data = request.get_json(force=True)
    sigma1 = float(data['sri'])         # 10³次循环下的应力（SRI）
    sigma2 = float(data['fatigue'])     # 10⁶次循环下的疲劳极限

    N1, N2 = 1e3, 1e6

    # Basquin参数 a, b
    b = (np.log10(sigma2) - np.log10(sigma1)) / (np.log10(N2) - np.log10(N1))
    log_a = np.log10(sigma1) - b * np.log10(N1)
    a = 10 ** log_a

    N_vals = np.logspace(3, 7, 300)  # 从10^3到10^7
    sigma_vals = a * N_vals ** b

    # ========== 图1：对数坐标图 ==========
    fig1, ax1 = plt.subplots()
    ax1.loglog(N_vals, sigma_vals, color='blue', label='预测 S-N 曲线')
    ax1.set_xlabel('循环次数 N（对数坐标）')
    ax1.set_ylabel('应力幅 σa / MPa')
    ax1.set_title('预测疲劳寿命曲线（对数坐标）')
    ax1.grid(True, which='both', linestyle='--', linewidth=0.5)
    ax1.legend()

    buf1 = io.BytesIO()
    plt.savefig(buf1, format='png')
    buf1.seek(0)
    log_img = base64.b64encode(buf1.read()).decode('utf-8')
    plt.close(fig1)

    # ========== 图2：线性坐标图 ==========
    fig2, ax2 = plt.subplots()
    ax2.plot(N_vals, sigma_vals, color='green', label='预测 S-N 曲线（线性）')
    ax2.set_xlabel('循环次数 N')
    ax2.set_ylabel('应力幅 σa / MPa')
    ax2.set_title('预测疲劳寿命曲线（线性坐标）')
    ax2.grid(True, linestyle='--', linewidth=0.5)
    ax2.legend()

    buf2 = io.BytesIO()
    plt.savefig(buf2, format='png')
    buf2.seek(0)
    linear_img = base64.b64encode(buf2.read()).decode('utf-8')
    plt.close(fig2)

    # 返回生成的图像数据
    return jsonify({
        'log_plot': log_img,
        'linear_plot': linear_img,
        'parameters': {
            'a': a,
            'b': b,
            'sigma1': sigma1,
            'sigma2': sigma2
        }
    })
# 批量预测处理接口
@app.route('/predict_batch', methods=['POST'])
def predict_batch():
    try:
        # 检查是否有文件上传
        if 'file' not in request.files:
            return jsonify({'error': '未找到上传的文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '未选择文件'}), 400
        
        # 验证文件类型
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'error': '请上传Excel文件（.xlsx或.xls格式）'}), 400
        
        # 读取Excel文件
        try:
            df = pd.read_excel(file)
        except Exception as e:
            return jsonify({'error': f'文件读取失败: {str(e)}'}), 400
        
        # 验证列数
        if len(df.columns) != 7:
            return jsonify({'error': f'Excel文件必须包含7列，当前有{len(df.columns)}列'}), 400
        
        # 提取所需的特征数据（按顺序）
        # 顺序：C, P, S, YS, UTS, Brinell, Elongation
        try:
            # 只需前5列用于预测：C, P, S, YS, UTS
            feature_data = pd.DataFrame({
                'C': df.iloc[:, 0],    # 第1列 C
                'P': df.iloc[:, 1],    # 第2列 P
                'S': df.iloc[:, 2],    # 第3列 S
                'YS': df.iloc[:, 3],   # 第4列 YS
                'UTS': df.iloc[:, 4]   # 第5列 UTS
            })
        except IndexError as e:
            return jsonify({'error': f'无法提取指定列的数据: {str(e)}'}), 400
        
        # 验证特征数据的数据类型
        required_columns = ['C', 'P', 'S', 'YS', 'UTS']
        for col in required_columns:
            if not pd.api.types.is_numeric_dtype(feature_data[col]):
                col_names = {
                    'C': '第1列(C)',
                    'P': '第2列(P)', 
                    'S': '第3列(S)',
                    'YS': '第4列(YS)',
                    'UTS': '第5列(UTS)'
                }
                return jsonify({'error': f'{col_names[col]} 必须为数字类型'}), 400
        
        # 检查特征数据是否有空值
        if feature_data.isnull().any().any():
            return jsonify({'error': '特征列中包含空值，请检查前5列的数据'}), 400
        
        # 限制数据行数
        if len(df) > 1000:
            return jsonify({'error': '数据行数不能超过1000行'}), 400
        
        # 进行批量预测
        try:
            # 使用现有模型进行预测
            fatigue_predictions = rf_fatigue.predict(feature_data)
            sri_predictions = rf_sri.predict(feature_data)
            
            # 创建结果DataFrame：保留原有7列 + 添加3列预测结果
            result_df = df.copy()
            result_df['疲劳极限'] = [round(x, 3) for x in fatigue_predictions]
            result_df['疲劳极限(log10)'] = [round(np.log10(x), 3) if x > 0 else None for x in fatigue_predictions]
            result_df['SRI(MPa)'] = [round(x, 3) for x in sri_predictions]
            
            # 保存结果到Excel文件
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                result_df.to_excel(writer, sheet_name='预测结果', index=False)
                
                # 获取工作表并设置列宽
                worksheet = writer.sheets['预测结果']
                
                # 设置列宽：原有7列 + 3列预测结果
                # 前7列设置为标准宽度
                for col_idx in range(1, 8):  # A-G列（原有数据）
                    col_letter = chr(65 + col_idx - 1)  # A=65, B=66, etc.
                    worksheet.column_dimensions[col_letter].width = 12
                
                # 预测结果列设置特殊宽度（第8、9、10列）
                worksheet.column_dimensions['H'].width = 15   # 疲劳极限
                worksheet.column_dimensions['I'].width = 18   # 疲劳极限(log10)
                worksheet.column_dimensions['J'].width = 15   # SRI(MPa)
                
                # 设置样式
                from openpyxl.styles import Font, Alignment, PatternFill
                
                # 定义字体和对齐方式
                title_font = Font(name='Calibri', bold=True, size=12)
                data_font = Font(name='Calibri', size=12)
                title_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
                center_alignment = Alignment(horizontal='center', vertical='center')
                
                # 应用标题样式到所有列（7列原数据 + 3列预测结果）
                for col in range(1, 11):  # 10列
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = title_font
                    # 预测结果列使用不同的背景色
                    if col > 7:  # 第8、9、10列是预测结果
                        cell.fill = PatternFill(start_color='FFE6CC', end_color='FFE6CC', fill_type='solid')
                    else:
                        cell.fill = title_fill
                    cell.alignment = center_alignment
                
                # 设置数据行的字体、对齐方式和数字格式
                for row in range(2, len(result_df) + 2):
                    for col in range(1, 11):  # 10列
                        cell = worksheet.cell(row=row, column=col)
                        cell.font = data_font  # 设置Calibri字体
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # 根据列号设置数字格式
                        if col in [4, 5, 7]:  # 第4,5,7列保留1位小数
                            cell.number_format = '0.0'
                        elif col in [6]:  # 第6列保留0位小数
                            cell.number_format = '0'
                        elif col in [1, 8, 9]:   # 第1,8,9列保留2位小数
                            cell.number_format = '0.00'
                        elif col in [2, 3, 10]:  # 第2,3,10列保留4位小数
                            cell.number_format = '0.0000'
                        # 其他列（预测结果列）保持默认3位小数
            
            output.seek(0)
            
            # 创建响应并添加样本数到header
            response = make_response(send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='批量预测结果.xlsx'
            ))
            response.headers['X-Total-Samples'] = str(len(result_df))
            return response
            
        except Exception as e:
            return jsonify({'error': f'预测过程中出现错误: {str(e)}'}), 500
            
    except Exception as e:
        return jsonify({'error': f'处理请求时出现错误: {str(e)}'}), 500

if __name__ == '__main__':
    # 支持云端部署的动态端口
    port = int(os.environ.get('PORT', 5000))
    # 绑定到0.0.0.0以便外部访问
    app.run(host='0.0.0.0', port=port, debug=False)